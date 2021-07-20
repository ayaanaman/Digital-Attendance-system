import cv2
import numpy as np
import face_recognition
import os
import openpyxl as op
from datetime import datetime
from tkinter import*

def getEncodings(images):
    encodeList=[]
    img = images
    img=cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
    encode=face_recognition.face_encodings(img)
    encodeList.append(encode)
    return encodeList

def atten_db(nameList,Date):
    wb = op.load_workbook("IT_OOPS_attendence.xlsx")
    sht = wb.get_sheet_by_name('IT_OOPS_attendence')
    j = sht.max_column + 1
    sht.cell(row=1, column=j).value = Date
    for i in range(2, sht.max_row + 1):
        if sht.cell(row=i, column=1).value in nameList:
            sht.cell(row=i, column=j).value = 'P'
        else:
            sht.cell(row=i, column=j).value = 'A'
    wb.save("IT_OOPS_attendence.xlsx")

def attendance(name,Date):
    with open("temp.csv",'r+')as f:
        dblist=f.readlines()
        nameList=[]
        for l in dblist:
            new=l.split(',')
            nameList.append(new[0])
        if name not in nameList:
            now=datetime.now()
            dt=now.strftime('%H:%M:%S')
            f.writelines(f'\n{name},{Date},{dt}')


def rem_stu(name):
    os.remove(f"students images/{name}.jpg")
    i=0
    while(imgNames[i].lower()<name.lower()):
        i+=1
    List = np.loadtxt("data_base.csv", delimiter=",")
    List=np.delete(List,i,0)
    f = open("data_base.csv", "wb")
    np.savetxt(f, List, delimiter=",")
    f.close()

    wb = op.load_workbook("IT_OOPS_attendence.xlsx")
    sht = wb.get_sheet_by_name('IT_OOPS_attendence')
    sht.delete_rows(i+2, amount=1)
    wb.save("IT_OOPS_attendence.xlsx")

def add_stu(name):
    cap = cv2.VideoCapture(0)
    i=0
    while(True):
        flag, add_img = cap.read()
        cv2.imshow("press C to capture",add_img)
        k=cv2.waitKey(1) & 0xFF
        if(k==ord('c') or k==ord('C')):
            while(imgNames[i].lower()<name.lower()):
                i += 1
            en = getEncodings(add_img)
            en = np.array(en)
            en = np.reshape(en, (1, np.product(en.shape)))

            List = np.loadtxt("data_base.csv", delimiter=",")
            List = np.insert(List, [i], en, axis=0)
            f = open("data_base.csv", "wb")
            np.savetxt(f, List, delimiter=",")
            f.close()
            cv2.imwrite(f"students images/{name}.jpg", add_img)

            break
    cap.release()
    cv2.destroyAllWindows()

    wb = op.load_workbook("IT_OOPS_attendence.xlsx")
    sht = wb.get_sheet_by_name('IT_OOPS_attendence')
    sht.insert_rows(i+2, amount=1)
    sht.cell(row=i+2, column=1).value = name.upper()
    for j in range(2, sht.max_column + 1):
        sht.cell(row=i+2, column=j).value = 'A'
    wb.save("IT_OOPS_attendence.xlsx")


def main(Date):
    cap = cv2.VideoCapture(0)
    encodeList = np.loadtxt("data_base.csv", delimiter=",")
    while True:
        flag, img = cap.read()
        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
        facesInCurrFrame = face_recognition.face_locations(imgS)
        encodeInCurrFrame = face_recognition.face_encodings(imgS, facesInCurrFrame)

        for camEncode, camFaceLoc in zip(encodeInCurrFrame, facesInCurrFrame):
            match = face_recognition.compare_faces(encodeList, camEncode)
            dis = face_recognition.face_distance(encodeList, camEncode)
            matIndx = np.argmin(dis)

            if match[matIndx]:
                name = imgNames[matIndx].upper()
                y1, x2, y2, x1 = camFaceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_PLAIN, 1, (255, 255, 255), 2)
                attendance(name,Date)
        cv2.imshow("camON", img)
        if cv2.waitKey(1) & 0xff == ord('q'):
            break
    cap.release()
    with open("temp.csv",'r')as f:
        dblist=f.readlines()
        nameList=[]
        for l in dblist:
            new=l.split(',')
            nameList.append(new[0])
    atten_db(nameList,Date)

def graph(name):
    pass

def quit():
    root.destroy()

def var(f):
    global flag
    flag=f

def fun(f,data):
    if (f == 1):
        main(data)
    elif (f == 2):
        graph(data)
    elif (f == 3):
        add_stu(data)
    elif (f == 4):
        rem_stu(data)

flag=0
path="students images"
imgNames=[]
imglist=os.listdir(path)

open('temp.csv', 'w').close()

for img in imglist:
    imgNames.append(os.path.splitext(img)[0])

root=Tk()
root.geometry("%dx%d" % (400+60,400))
photo = PhotoImage(file="icons/icon.png")
root.iconphoto(False, photo)
root.title(" Attendence System ")
frame=Frame().pack(side=TOP,fill=BOTH)
frame1=Frame().pack(side=BOTTOM,fill=BOTH)
Label(frame,text=" Digital Attendence System ",font=("Comicsansms",20,"bold"),fg="royalblue3").pack(fill=BOTH)
photo=PhotoImage(file="icons/attendence.png")
photo=photo.subsample(3,3)
photo2=PhotoImage(file="icons/data.png")
photo2=photo2.subsample(3,3)
photo3=PhotoImage(file="icons/add.png")
photo3=photo3.subsample(3,3)
photo4=PhotoImage(file="icons/delete.png")
photo4=photo4.subsample(3,3)
Button(frame,text="Attendance",font=("Arial", 10,"bold"), image=photo, width=100,height=100,compound=TOP,fg="snow",bg="black",relief=RAISED,command=lambda:[var(1),quit()]).pack(side=LEFT,padx=40)
Button(frame,text="Analysis",font=("Arial", 10,"bold"), image=photo2, width=100,height=100,compound=TOP,fg="snow",bg="black",relief=RAISED,command=lambda:[var(2),quit()]).pack(side=RIGHT,padx=40)
Button(frame1,text="Add Std",font=("Arial", 10,"bold"), image=photo3, width=80,height=80,compound=TOP,fg="snow",bg="black",relief=RAISED,command=lambda:[var(3),quit()]).pack(pady=40)
Button(frame1,text="Delete Std",font=("Arial", 10,"bold"), image=photo4, width=80,height=80,compound=TOP,fg="snow",bg="black",relief=RAISED,command=lambda:[var(4),quit()]).pack(pady=40)
root.mainloop()

if(flag==1):
    path="Enter Date (dd.mm.yy): "
    ins='''*press 'Q' when done!!! 
           *You can view updated attendence in IT_OOPS_attendence file'''
elif(flag in [2,3,4]):
    path="Enter Name: "
    if(flag==3):
       ins="*press 'C' to capture!!!"
    else:
        ins=" "

root = Tk()
root.geometry("600x200")
n1 = StringVar()
Label(root, text=path, font=("Arial", 13, "bold")).grid(row=3, column=1)
Entry(root, textvariable=n1).grid(row=3, column=2)
Button(root, text="Submit", command=lambda:[fun(flag,n1.get()),quit()]).grid(row=5, column=2)
Label(root, text=ins, font=("Arial", 8, "bold")).grid(row=6, column=1)
root.mainloop()

